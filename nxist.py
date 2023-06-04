import json
from openpyxl import Workbook

teams = json.load(open("teams.json", "r", encoding="UTF-8"))
organizations = json.load(open("organizations.json", "r", encoding="UTF-8"))
submissions = json.load(open("submissions.json", "r", encoding="UTF-8"))

od = {}
for org in organizations:
    od[org["id"]] = org["name"]

td = {}
for team in teams:
    td[team["id"]] = [od[team["organization_id"]], team["name"], False]

for sub in submissions:
    td[sub["team_id"]][2] = True

book = Workbook()
sheet = book.active
for i, (id, ls) in enumerate(td.items()):
    sheet.cell(i + 1, 1).value = ls[0]
    sheet.cell(i + 1, 2).value = ls[1]
    sheet.cell(i + 1, 3).value = ls[2]

book.save("nxist.xlsx")
